# data_import_export.py

import json
import csv
import traceback
import os
import shutil
import zipfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

# Conditional imports with fallbacks
try:
    from system_logger import LoggerFactory
    logger = LoggerFactory.get_logger('data_import_export')
except ImportError:
    import logging
    logger = logging.getLogger('data_import_export')
    logging.basicConfig(level=logging.INFO)

class DataImportExport:
    """Comprehensive data import/export system for the port management application."""
    
    def __init__(self, database_connection):
        self.db = database_connection
        self.backup_dir = "backups"
        self.temp_dir = "temp"
        self.ensure_directories()

    def ensure_directories(self):
        """Ensure required directories exist."""
        for directory in [self.backup_dir, self.temp_dir]:
            if not os.path.exists(directory):
                os.makedirs(directory)

    # EXCEL IMPORT/EXPORT METHODS
    
    def export_to_excel(self, output_path: str, include_tables: List[str] = None) -> bool:
        """Export database data to Excel format."""
        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define tables to export
            if include_tables is None:
                include_tables = ['containers', 'ships', 'container_lifecycle', 'notifications']
            
            for table_name in include_tables:
                try:
                    # Get table data
                    cursor = self.db.connection.cursor()
                    cursor.execute(f"SELECT * FROM {table_name}")
                    data = cursor.fetchall()
                    
                    # Get column names
                    cursor.execute(f"""
                        SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name = '{table_name}'
                        ORDER BY ordinal_position
                    """)
                    columns = [row[0] for row in cursor.fetchall()]
                    cursor.close()
                    
                    if data:
                        # Create worksheet
                        ws = wb.create_sheet(title=table_name.title())
                        
                        # Add headers
                        for col, header in enumerate(columns, 1):
                            cell = ws.cell(row=1, column=col, value=header.title())
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center")
                        
                        # Add data
                        for row_idx, row_data in enumerate(data, 2):
                            for col_idx, value in enumerate(row_data, 1):
                                # Handle datetime objects
                                if isinstance(value, datetime):
                                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                                ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        logger.info(f"Exported {len(data)} records from {table_name}")
                    else:
                        logger.warning(f"No data found in table {table_name}")
                        
                except Exception as e:
                    logger.error(f"Error exporting table {table_name}: {str(e)}")
                    continue
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel export completed successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error during Excel export: {str(e)}")
            logger.error(traceback.format_exc())
            return False

    def import_from_excel(self, file_path: str, table_mappings: Dict[str, str] = None) -> Dict[str, Any]:
        """Import data from Excel file."""
        try:
            results = {
                'success': False,
                'imported_tables': {},
                'errors': [],
                'warnings': []
            }
            
            wb = load_workbook(file_path, data_only=True)
            
            # Default table mappings (worksheet name -> table name)
            if table_mappings is None:
                table_mappings = {
                    'Containers': 'containers',
                    'Ships': 'ships',
                    'Container_Lifecycle': 'container_lifecycle',
                    'Notifications': 'notifications'
                }
            
            for sheet_name in wb.sheetnames:
                if sheet_name in table_mappings:
                    table_name = table_mappings[sheet_name]
                    try:
                        ws = wb[sheet_name]
                        imported_count = self._import_worksheet_data(ws, table_name)
                        results['imported_tables'][table_name] = imported_count
                        logger.info(f"Imported {imported_count} records into {table_name}")
                    except Exception as e:
                        error_msg = f"Error importing sheet {sheet_name}: {str(e)}"
                        results['errors'].append(error_msg)
                        logger.error(error_msg)
                else:
                    results['warnings'].append(f"Skipped unknown sheet: {sheet_name}")
            
            results['success'] = len(results['imported_tables']) > 0
            return results
            
        except Exception as e:
            logger.error(f"Error during Excel import: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'success': False,
                'imported_tables': {},
                'errors': [str(e)],
                'warnings': []
            }

    def _import_worksheet_data(self, worksheet, table_name: str) -> int:
        """Import data from a specific worksheet."""
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value.lower().replace(' ', '_'))
        
        if not headers:
            raise ValueError("No headers found in worksheet")
        
        # Get table columns to validate against
        cursor = self.db.connection.cursor()
        cursor.execute(f"""
            SELECT column_name, data_type 
            FROM information_schema.columns 
            WHERE table_name = '{table_name}'
        """)
        table_columns = {row[0]: row[1] for row in cursor.fetchall()}
        
        # Validate headers
        valid_headers = []
        for header in headers:
            if header in table_columns:
                valid_headers.append(header)
            else:
                logger.warning(f"Column {header} not found in table {table_name}")
        
        if not valid_headers:
            raise ValueError("No valid columns found for import")
        
        # Import data
        imported_count = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
                
            try:
                # Build insert data
                insert_data = {}
                for i, header in enumerate(valid_headers):
                    if i < len(row) and row[i] is not None:
                        value = row[i]
                        # Handle different data types
                        if isinstance(value, str) and value.strip() == '':
                            value = None
                        insert_data[header] = value
                
                if insert_data:
                    # Generate INSERT query
                    columns = list(insert_data.keys())
                    placeholders = ', '.join(['%s'] * len(columns))
                    column_names = ', '.join(columns)
                    
                    insert_query = f"""
                        INSERT INTO {table_name} ({column_names}) 
                        VALUES ({placeholders})
                        ON CONFLICT DO NOTHING
                    """
                    
                    cursor.execute(insert_query, list(insert_data.values()))
                    imported_count += 1
                    
            except Exception as e:
                logger.error(f"Error importing row: {str(e)}")
                continue
        
        self.db.connection.commit()
        cursor.close()
        return imported_count

    # CSV IMPORT/EXPORT METHODS
    
    def export_to_csv(self, table_name: str, output_path: str, delimiter: str = ',') -> bool:
        """Export table data to CSV format."""
        try:
            cursor = self.db.connection.cursor()
            cursor.execute(f"SELECT * FROM {table_name}")
            data = cursor.fetchall()
            
            # Get column names
            cursor.execute(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}'
                ORDER BY ordinal_position
            """)
            columns = [row[0] for row in cursor.fetchall()]
            cursor.close()
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=delimiter)
                
                # Write headers
                writer.writerow(columns)
                
                # Write data
                for row in data:
                    # Convert datetime objects to strings
                    processed_row = []
                    for value in row:
                        if isinstance(value, datetime):
                            processed_row.append(value.strftime("%Y-%m-%d %H:%M:%S"))
                        else:
                            processed_row.append(value)
                    writer.writerow(processed_row)
            
            logger.info(f"CSV export completed: {output_path} ({len(data)} records)")
            return True
            
        except Exception as e:
            logger.error(f"Error during CSV export: {str(e)}")
            return False

    def import_from_csv(self, file_path: str, table_name: str, delimiter: str = ',') -> Dict[str, Any]:
        """Import data from CSV file."""
        try:
            result = {
                'success': False,
                'imported_count': 0,
                'errors': [],
                'skipped_rows': 0
            }
            
            # Read CSV file
            df = pd.read_csv(file_path, delimiter=delimiter)
            
            # Get table columns
            cursor = self.db.connection.cursor()
            cursor.execute(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}'
            """)
            table_columns = [row[0] for row in cursor.fetchall()]
            
            # Validate CSV columns
            csv_columns = [col.lower().replace(' ', '_') for col in df.columns]
            valid_columns = [col for col in csv_columns if col in table_columns]
            
            if not valid_columns:
                raise ValueError("No valid columns found for import")
            
            # Import data
            for index, row in df.iterrows():
                try:
                    insert_data = {}
                    for col in valid_columns:
                        value = row[df.columns[csv_columns.index(col)]]
                        if pd.isna(value):
                            value = None
                        insert_data[col] = value
                    
                    if insert_data:
                        columns = list(insert_data.keys())
                        placeholders = ', '.join(['%s'] * len(columns))
                        column_names = ', '.join(columns)
                        
                        insert_query = f"""
                            INSERT INTO {table_name} ({column_names}) 
                            VALUES ({placeholders})
                            ON CONFLICT DO NOTHING
                        """
                        
                        cursor.execute(insert_query, list(insert_data.values()))
                        result['imported_count'] += 1
                        
                except Exception as e:
                    result['errors'].append(f"Row {index + 1}: {str(e)}")
                    result['skipped_rows'] += 1
                    continue
            
            self.db.connection.commit()
            cursor.close()
            
            result['success'] = result['imported_count'] > 0
            logger.info(f"CSV import completed: {result['imported_count']} records imported")
            return result
            
        except Exception as e:
            logger.error(f"Error during CSV import: {str(e)}")
            return {
                'success': False,
                'imported_count': 0,
                'errors': [str(e)],
                'skipped_rows': 0
            }

    # DATABASE BACKUP/RESTORE METHODS
    
    def create_full_backup(self, backup_name: str = None) -> Dict[str, Any]:
        """Create a complete database backup."""
        try:
            if backup_name is None:
                backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            backup_path = os.path.join(self.backup_dir, f"{backup_name}.zip")
            temp_backup_dir = os.path.join(self.temp_dir, backup_name)
            
            # Create temp directory
            if os.path.exists(temp_backup_dir):
                shutil.rmtree(temp_backup_dir)
            os.makedirs(temp_backup_dir)
            
            # Get all tables
            cursor = self.db.connection.cursor()
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = 'public'
            """)
            tables = [row[0] for row in cursor.fetchall()]
            cursor.close()
            
            backup_info = {
                'backup_name': backup_name,
                'created_at': datetime.now().isoformat(),
                'tables': {},
                'total_records': 0
            }
            
            # Export each table
            for table_name in tables:
                try:
                    table_file = os.path.join(temp_backup_dir, f"{table_name}.json")
                    record_count = self._export_table_to_json(table_name, table_file)
                    backup_info['tables'][table_name] = record_count
                    backup_info['total_records'] += record_count
                    logger.info(f"Backed up {record_count} records from {table_name}")
                except Exception as e:
                    logger.error(f"Error backing up table {table_name}: {str(e)}")
                    backup_info['tables'][table_name] = f"ERROR: {str(e)}"
            
            # Save backup info
            info_file = os.path.join(temp_backup_dir, "backup_info.json")
            with open(info_file, 'w') as f:
                json.dump(backup_info, f, indent=2)
            
            # Create zip file
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(temp_backup_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, temp_backup_dir)
                        zipf.write(file_path, arcname)
            
            # Cleanup temp directory
            shutil.rmtree(temp_backup_dir)
            
            logger.info(f"Full backup created successfully: {backup_path}")
            return {
                'success': True,
                'backup_path': backup_path,
                'backup_info': backup_info
            }
            
        except Exception as e:
            logger.error(f"Error creating backup: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'success': False,
                'error': str(e)
            }

    def restore_from_backup(self, backup_path: str, restore_options: Dict[str, Any] = None) -> Dict[str, Any]:
        """Restore database from backup file."""
        try:
            if restore_options is None:
                restore_options = {
                    'clear_existing': False,
                    'tables_to_restore': None,  # None means all tables
                    'backup_existing': True
                }
            
            # Create backup of current data if requested
            if restore_options.get('backup_existing', True):
                current_backup = self.create_full_backup("pre_restore_backup")
                if not current_backup['success']:
                    logger.warning("Could not create pre-restore backup")
            
            # Extract backup
            temp_restore_dir = os.path.join(self.temp_dir, "restore_temp")
            if os.path.exists(temp_restore_dir):
                shutil.rmtree(temp_restore_dir)
            os.makedirs(temp_restore_dir)
            
            with zipfile.ZipFile(backup_path, 'r') as zipf:
                zipf.extractall(temp_restore_dir)
            
            # Read backup info
            info_file = os.path.join(temp_restore_dir, "backup_info.json")
            if not os.path.exists(info_file):
                raise ValueError("Invalid backup file: missing backup_info.json")
            
            with open(info_file, 'r') as f:
                backup_info = json.load(f)
            
            # Determine tables to restore
            tables_to_restore = restore_options.get('tables_to_restore')
            if tables_to_restore is None:
                tables_to_restore = list(backup_info['tables'].keys())
            
            restore_results = {
                'success': False,
                'restored_tables': {},
                'errors': [],
                'total_restored': 0
            }
            
            # Restore each table
            for table_name in tables_to_restore:
                if table_name not in backup_info['tables']:
                    restore_results['errors'].append(f"Table {table_name} not found in backup")
                    continue
                
                table_file = os.path.join(temp_restore_dir, f"{table_name}.json")
                if not os.path.exists(table_file):
                    restore_results['errors'].append(f"Data file for {table_name} not found")
                    continue
                
                try:
                    # Clear existing data if requested
                    if restore_options.get('clear_existing', False):
                        cursor = self.db.connection.cursor()
                        cursor.execute(f"DELETE FROM {table_name}")
                        cursor.close()
                        logger.info(f"Cleared existing data from {table_name}")
                    
                    # Restore table data
                    restored_count = self._import_table_from_json(table_name, table_file)
                    restore_results['restored_tables'][table_name] = restored_count
                    restore_results['total_restored'] += restored_count
                    logger.info(f"Restored {restored_count} records to {table_name}")
                    
                except Exception as e:
                    error_msg = f"Error restoring {table_name}: {str(e)}"
                    restore_results['errors'].append(error_msg)
                    logger.error(error_msg)
            
            # Cleanup temp directory
            shutil.rmtree(temp_restore_dir)
            
            restore_results['success'] = len(restore_results['restored_tables']) > 0
            logger.info(f"Restore completed: {restore_results['total_restored']} total records restored")
            return restore_results
            
        except Exception as e:
            logger.error(f"Error during restore: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'success': False,
                'error': str(e),
                'restored_tables': {},
                'errors': [str(e)],
                'total_restored': 0
            }

    def _export_table_to_json(self, table_name: str, output_file: str) -> int:
        """Export table data to JSON file."""
        cursor = self.db.connection.cursor()
        cursor.execute(f"SELECT * FROM {table_name}")
        data = cursor.fetchall()
        
        # Get column names
        cursor.execute(f"""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = '{table_name}'
            ORDER BY ordinal_position
        """)
        columns = [row[0] for row in cursor.fetchall()]
        cursor.close()
        
        # Convert to list of dictionaries
        records = []
        for row in data:
            record = {}
            for i, value in enumerate(row):
                if isinstance(value, datetime):
                    value = value.isoformat()
                record[columns[i]] = value
            records.append(record)
        
        # Save to JSON
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(records, f, indent=2, ensure_ascii=False)
        
        return len(records)

    def _import_table_from_json(self, table_name: str, json_file: str) -> int:
        """Import table data from JSON file."""
        with open(json_file, 'r', encoding='utf-8') as f:
            records = json.load(f)
        
        if not records:
            return 0
        
        cursor = self.db.connection.cursor()
        imported_count = 0
        
        for record in records:
            try:
                columns = list(record.keys())
                values = list(record.values())
                
                # Handle datetime strings
                for i, value in enumerate(values):
                    if isinstance(value, str) and 'T' in value and ':' in value:
                        try:
                            values[i] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        except:
                            pass  # Keep as string if conversion fails
                
                placeholders = ', '.join(['%s'] * len(columns))
                column_names = ', '.join(columns)
                
                insert_query = f"""
                    INSERT INTO {table_name} ({column_names}) 
                    VALUES ({placeholders})
                    ON CONFLICT DO NOTHING
                """
                
                cursor.execute(insert_query, values)
                imported_count += 1
                
            except Exception as e:
                logger.error(f"Error importing record to {table_name}: {str(e)}")
                continue
        
        self.db.connection.commit()
        cursor.close()
        return imported_count

    # UTILITY METHODS
    
    def get_backup_list(self) -> List[Dict[str, Any]]:
        """Get list of available backups."""
        backups = []
        
        if not os.path.exists(self.backup_dir):
            return backups
        
        for filename in os.listdir(self.backup_dir):
            if filename.endswith('.zip'):
                filepath = os.path.join(self.backup_dir, filename)
                try:
                    stat = os.stat(filepath)
                    backup_info = {
                        'name': filename[:-4],  # Remove .zip extension
                        'file_path': filepath,
                        'size': stat.st_size,
                        'created_at': datetime.fromtimestamp(stat.st_mtime)
                    }
                    
                    # Try to read backup info from zip
                    try:
                        with zipfile.ZipFile(filepath, 'r') as zipf:
                            if 'backup_info.json' in zipf.namelist():
                                info_data = zipf.read('backup_info.json')
                                info = json.loads(info_data.decode('utf-8'))
                                # Convert string dates back to datetime objects
                                if 'created_at' in info and isinstance(info['created_at'], str):
                                    try:
                                        info['created_at'] = datetime.fromisoformat(info['created_at'].replace('Z', '+00:00'))
                                    except ValueError:
                                        # Keep the original datetime if parsing fails
                                        pass
                                backup_info.update(info)
                    except:
                        pass  # Use basic info if can't read details
                    
                    backups.append(backup_info)
                except Exception as e:
                    logger.error(f"Error reading backup info for {filename}: {str(e)}")
        
        return sorted(backups, key=lambda x: x['created_at'], reverse=True)

    def delete_backup(self, backup_name: str) -> bool:
        """Delete a backup file."""
        try:
            backup_path = os.path.join(self.backup_dir, f"{backup_name}.zip")
            if os.path.exists(backup_path):
                os.remove(backup_path)
                logger.info(f"Deleted backup: {backup_name}")
                return True
            else:
                logger.warning(f"Backup not found: {backup_name}")
                return False
        except Exception as e:
            logger.error(f"Error deleting backup {backup_name}: {str(e)}")
            return False

    def validate_data_integrity(self, table_name: str) -> Dict[str, Any]:
        """Validate data integrity for a table."""
        try:
            cursor = self.db.connection.cursor()
            
            # Basic statistics
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            total_records = cursor.fetchone()[0]
            
            # Check for null values in required fields
            cursor.execute(f"""
                SELECT column_name, is_nullable 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}'
            """)
            columns_info = cursor.fetchall()
            
            null_checks = {}
            for column, is_nullable in columns_info:
                if is_nullable == 'NO':
                    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE {column} IS NULL")
                    null_count = cursor.fetchone()[0]
                    if null_count > 0:
                        null_checks[column] = null_count
            
            cursor.close()
            
            return {
                'table_name': table_name,
                'total_records': total_records,
                'null_violations': null_checks,
                'integrity_score': 100 - (len(null_checks) * 10),  # Simple scoring
                'valid': len(null_checks) == 0
            }
            
        except Exception as e:
            logger.error(f"Error validating data integrity for {table_name}: {str(e)}")
            return {
                'table_name': table_name,
                'error': str(e),
                'valid': False
            }
